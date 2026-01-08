"""
Router para generación de reportes ejecutivos
Soporta múltiples formatos: JSON, CSV, Excel, PDF
"""

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from datetime import datetime, timedelta
from typing import Optional, List
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from auth import get_current_user
from db import get_db_connection_citas
from reportes.pdf_generator import generar_pdf_gastos, generar_pdf_inventario

router = APIRouter(prefix="/reportes", tags=["Reportes"])


@router.get("/gastos-mensuales")
async def generar_reporte_gastos_mensuales(
    mes: int = Query(..., ge=1, le=12, description="Mes del reporte (1-12)"),
    anio: int = Query(..., ge=2020, description="Año del reporte"),
    formato: str = Query("json", pattern="^(json|csv|excel|pdf)$", description="Formato de exportación"),
    current_user: dict = Depends(get_current_user)
):
    """
    Genera reporte consolidado de gastos mensuales con:
    - Resumen por categoría
    - Comparativa con mes anterior
    - Tendencia de 6 meses
    - Top 10 gastos mayores
    - Productos comprados (si vinculados)
    """
    conn = await get_db_connection_citas()
    
    try:
        # Calcular fechas
        primer_dia = datetime(anio, mes, 1)
        if mes == 12:
            ultimo_dia = datetime(anio + 1, 1, 1) - timedelta(days=1)
        else:
            ultimo_dia = datetime(anio, mes + 1, 1) - timedelta(days=1)
        
        # Mes anterior para comparativa
        if mes == 1:
            mes_anterior = 12
            anio_anterior = anio - 1
        else:
            mes_anterior = mes - 1
            anio_anterior = anio
        
        primer_dia_anterior = datetime(anio_anterior, mes_anterior, 1)
        if mes_anterior == 12:
            ultimo_dia_anterior = datetime(anio_anterior + 1, 1, 1) - timedelta(days=1)
        else:
            ultimo_dia_anterior = datetime(anio_anterior, mes_anterior + 1, 1) - timedelta(days=1)
        
        # Query para gastos del mes actual
        gastos_mes = await conn.fetch("""
            SELECT 
                categoria,
                SUM(monto) as total,
                COUNT(*) as cantidad
            FROM gastos
            WHERE fecha_gasto BETWEEN $1 AND $2
            GROUP BY categoria
            ORDER BY total DESC
        """, primer_dia, ultimo_dia)
        
        # Query para gastos del mes anterior
        gastos_mes_anterior = await conn.fetch("""
            SELECT 
                categoria,
                SUM(monto) as total
            FROM gastos
            WHERE fecha_gasto BETWEEN $1 AND $2
            GROUP BY categoria
        """, primer_dia_anterior, ultimo_dia_anterior)
        
        # Crear diccionario para comparativa
        gastos_anterior_dict = {g['categoria']: float(g['total']) for g in gastos_mes_anterior}
        
        # Calcular totales y variaciones
        total_gastos = sum(float(g['total']) for g in gastos_mes)
        total_gastos_anterior = sum(gastos_anterior_dict.values())
        
        gastos_por_categoria = []
        for g in gastos_mes:
            categoria = g['categoria'] or 'SIN_CATEGORIA'
            total = float(g['total'])
            porcentaje = (total / total_gastos * 100) if total_gastos > 0 else 0
            
            total_anterior = gastos_anterior_dict.get(categoria, 0)
            if total_anterior > 0:
                variacion = ((total - total_anterior) / total_anterior) * 100
            else:
                variacion = 100 if total > 0 else 0
            
            gastos_por_categoria.append({
                'categoria': categoria,
                'total': total,
                'cantidad': g['cantidad'],
                'porcentaje': round(porcentaje, 2),
                'variacion_mes_anterior': round(variacion, 2)
            })
        
        # Top 10 gastos mayores
        top_gastos = await conn.fetch("""
            SELECT 
                concepto,
                monto,
                fecha_gasto,
                metodo_pago,
                categoria
            FROM gastos
            WHERE fecha_gasto BETWEEN $1 AND $2
            ORDER BY monto DESC
            LIMIT 10
        """, primer_dia, ultimo_dia)
        
        # Tendencia de 6 meses
        fecha_inicio_tendencia = primer_dia - timedelta(days=180)
        tendencia = await conn.fetch("""
            SELECT 
                EXTRACT(YEAR FROM fecha_gasto) as anio,
                EXTRACT(MONTH FROM fecha_gasto) as mes,
                SUM(monto) as total
            FROM gastos
            WHERE fecha_gasto BETWEEN $1 AND $2
            GROUP BY anio, mes
            ORDER BY anio, mes
        """, fecha_inicio_tendencia, ultimo_dia)
        
        # Productos comprados (vinculados a gastos)
        productos_comprados = await conn.fetch("""
            SELECT 
                p.nombre as producto,
                gi.cantidad_comprada,
                gi.precio_unitario,
                gi.cantidad_comprada * gi.precio_unitario as subtotal,
                g.concepto,
                g.fecha_gasto
            FROM gastos_inventario gi
            JOIN inventario_productos p ON gi.producto_id = p.producto_id
            JOIN gastos g ON gi.gasto_id = g.gasto_id
            WHERE g.fecha_gasto BETWEEN $1 AND $2
            ORDER BY gi.fecha_entrada DESC
        """, primer_dia, ultimo_dia)
        
        # Nombre del mes en español
        meses_es = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        periodo_nombre = f"{meses_es[mes - 1]} {anio}"
        
        # Construir reporte
        reporte = {
            'periodo': periodo_nombre,
            'fecha_inicio': primer_dia.strftime('%Y-%m-%d'),
            'fecha_fin': ultimo_dia.strftime('%Y-%m-%d'),
            'total_gastos': round(total_gastos, 2),
            'total_gastos_mes_anterior': round(total_gastos_anterior, 2),
            'variacion_porcentual': round(((total_gastos - total_gastos_anterior) / total_gastos_anterior * 100) if total_gastos_anterior > 0 else 0, 2),
            'gastos_por_categoria': gastos_por_categoria,
            'top_10_gastos': [
                {
                    'concepto': g['concepto'],
                    'monto': float(g['monto']),
                    'fecha': g['fecha_gasto'].strftime('%Y-%m-%d'),
                    'metodo_pago': g['metodo_pago'],
                    'categoria': g['categoria']
                }
                for g in top_gastos
            ],
            'tendencia_6_meses': [
                {
                    'periodo': f"{int(t['anio'])}-{int(t['mes']):02d}",
                    'total': float(t['total'])
                }
                for t in tendencia
            ],
            'productos_comprados': [
                {
                    'producto': p['producto'],
                    'cantidad': float(p['cantidad_comprada']),
                    'precio_unitario': float(p['precio_unitario']),
                    'subtotal': float(p['subtotal']),
                    'concepto_gasto': p['concepto'],
                    'fecha': p['fecha_gasto'].strftime('%Y-%m-%d')
                }
                for p in productos_comprados
            ]
        }
        
        # Generar según formato solicitado
        if formato == "csv":
            return generar_csv_gastos(reporte)
        elif formato == "excel":
            return generar_excel_gastos(reporte)
        elif formato == "pdf":
            pdf_buffer = generar_pdf_gastos(reporte)
            return StreamingResponse(
                pdf_buffer,
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f"attachment; filename=reporte_gastos_{reporte['periodo'].replace(' ', '_')}.pdf"
                }
            )
        else:
            return reporte
            
    finally:
        await conn.close()


def generar_csv_gastos(reporte: dict) -> StreamingResponse:
    """Genera CSV del reporte de gastos"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(['REPORTE DE GASTOS MENSUALES'])
    writer.writerow([f"Periodo: {reporte['periodo']}"])
    writer.writerow([f"Total Gastos: ${reporte['total_gastos']:,.2f}"])
    writer.writerow([f"Variación vs Mes Anterior: {reporte['variacion_porcentual']:+.2f}%"])
    writer.writerow([])
    
    # Gastos por categoría
    writer.writerow(['GASTOS POR CATEGORÍA'])
    writer.writerow(['Categoría', 'Total', 'Cantidad', 'Porcentaje', 'Variación %'])
    for cat in reporte['gastos_por_categoria']:
        writer.writerow([
            cat['categoria'],
            f"${cat['total']:,.2f}",
            cat['cantidad'],
            f"{cat['porcentaje']:.2f}%",
            f"{cat['variacion_mes_anterior']:+.2f}%"
        ])
    writer.writerow([])
    
    # Top 10 gastos
    writer.writerow(['TOP 10 GASTOS MAYORES'])
    writer.writerow(['Concepto', 'Monto', 'Fecha', 'Método Pago', 'Categoría'])
    for g in reporte['top_10_gastos']:
        writer.writerow([
            g['concepto'],
            f"${g['monto']:,.2f}",
            g['fecha'],
            g['metodo_pago'],
            g['categoria']
        ])
    writer.writerow([])
    
    # Productos comprados
    if reporte['productos_comprados']:
        writer.writerow(['PRODUCTOS COMPRADOS'])
        writer.writerow(['Producto', 'Cantidad', 'Precio Unitario', 'Subtotal', 'Concepto', 'Fecha'])
        for p in reporte['productos_comprados']:
            writer.writerow([
                p['producto'],
                p['cantidad'],
                f"${p['precio_unitario']:,.2f}",
                f"${p['subtotal']:,.2f}",
                p['concepto_gasto'],
                p['fecha']
            ])
    
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=reporte_gastos_{reporte['periodo'].replace(' ', '_')}.csv"
        }
    )


def generar_excel_gastos(reporte: dict) -> StreamingResponse:
    """Genera Excel del reporte de gastos"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Gastos"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    
    # Título
    ws['A1'] = 'REPORTE DE GASTOS MENSUALES'
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    
    ws['A2'] = f"Periodo: {reporte['periodo']}"
    ws['A3'] = f"Total Gastos: ${reporte['total_gastos']:,.2f}"
    ws['A4'] = f"Variación vs Mes Anterior: {reporte['variacion_porcentual']:+.2f}%"
    
    # Gastos por categoría
    row = 6
    ws[f'A{row}'] = 'GASTOS POR CATEGORÍA'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    headers = ['Categoría', 'Total', 'Cantidad', 'Porcentaje', 'Variación %']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    for cat in reporte['gastos_por_categoria']:
        ws.cell(row=row, column=1, value=cat['categoria'])
        ws.cell(row=row, column=2, value=cat['total']).number_format = '$#,##0.00'
        ws.cell(row=row, column=3, value=cat['cantidad'])
        ws.cell(row=row, column=4, value=cat['porcentaje']/100).number_format = '0.00%'
        ws.cell(row=row, column=5, value=cat['variacion_mes_anterior']/100).number_format = '+0.00%;-0.00%'
        row += 1
    
    # Top 10 gastos
    row += 2
    ws[f'A{row}'] = 'TOP 10 GASTOS MAYORES'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    headers = ['Concepto', 'Monto', 'Fecha', 'Método Pago', 'Categoría']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row += 1
    for g in reporte['top_10_gastos']:
        ws.cell(row=row, column=1, value=g['concepto'])
        ws.cell(row=row, column=2, value=g['monto']).number_format = '$#,##0.00'
        ws.cell(row=row, column=3, value=g['fecha'])
        ws.cell(row=row, column=4, value=g['metodo_pago'])
        ws.cell(row=row, column=5, value=g['categoria'])
        row += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Guardar en BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=reporte_gastos_{reporte['periodo'].replace(' ', '_')}.xlsx"
        }
    )


@router.get("/inventario-estado")
async def generar_reporte_inventario(
    formato: str = Query("json", pattern="^(json|csv|excel|pdf)$"),
    incluir_criticos: bool = Query(True),
    incluir_obsoletos: bool = Query(False),
    current_user: dict = Depends(get_current_user)
):
    """
    Genera reporte del estado actual del inventario con:
    - Valor total del inventario
    - Productos críticos (stock < mínimo * 1.3)
    - Productos con exceso (stock > máximo)
    - Rotación estimada
    - Productos sin movimiento en 90 días
    """
    conn = await get_db_connection_citas()
    
    try:
        # Valor total del inventario
        valor_total = await conn.fetchval("""
            SELECT COALESCE(SUM(stock_actual * COALESCE(costo_unitario, 0)), 0)
            FROM inventario_productos
            WHERE activo = TRUE
        """)
        
        # Número de productos
        num_productos = await conn.fetchval("""
            SELECT COUNT(*)
            FROM inventario_productos
            WHERE activo = TRUE
        """)
        
        # Productos críticos
        productos_criticos = []
        if incluir_criticos:
            productos_criticos = await conn.fetch("""
                SELECT 
                    producto_id,
                    codigo_producto,
                    nombre,
                    categoria,
                    stock_actual,
                    stock_minimo,
                    stock_maximo,
                    unidad_medida,
                    costo_unitario,
                    stock_actual * COALESCE(costo_unitario, 0) as valor_stock
                FROM inventario_productos
                WHERE activo = TRUE 
                AND stock_actual < (stock_minimo * 1.3)
                ORDER BY (stock_actual - stock_minimo) ASC
            """)
        
        # Productos con exceso
        productos_exceso = await conn.fetch("""
            SELECT 
                producto_id,
                codigo_producto,
                nombre,
                stock_actual,
                stock_maximo,
                unidad_medida
            FROM inventario_productos
            WHERE activo = TRUE 
            AND stock_actual > stock_maximo
            ORDER BY (stock_actual - stock_maximo) DESC
        """)
        
        # Productos sin movimiento en 90 días (si se solicita)
        productos_sin_movimiento = []
        if incluir_obsoletos:
            productos_sin_movimiento = await conn.fetch("""
                SELECT 
                    p.producto_id,
                    p.codigo_producto,
                    p.nombre,
                    p.stock_actual,
                    p.unidad_medida,
                    MAX(sm.fecha_movimiento) as ultimo_movimiento
                FROM inventario_productos p
                LEFT JOIN stock_movements sm ON p.producto_id = sm.producto_id
                WHERE p.activo = TRUE
                GROUP BY p.producto_id, p.codigo_producto, p.nombre, p.stock_actual, p.unidad_medida
                HAVING MAX(sm.fecha_movimiento) < NOW() - INTERVAL '90 days'
                OR MAX(sm.fecha_movimiento) IS NULL
                ORDER BY MAX(sm.fecha_movimiento) ASC NULLS FIRST
            """)
        
        # Rotación promedio (días)
        rotacion_promedio = await conn.fetchval("""
            SELECT AVG(tiempo_reposicion_dias)
            FROM inventario_productos
            WHERE activo = TRUE AND tiempo_reposicion_dias > 0
        """) or 0
        
        # Construir reporte
        reporte = {
            'fecha_generacion': datetime.now().isoformat(),
            'valor_total_inventario': round(float(valor_total), 2),
            'numero_productos': num_productos,
            'productos_criticos': [
                {
                    'producto_id': p['producto_id'],
                    'codigo': p['codigo_producto'],
                    'nombre': p['nombre'],
                    'categoria': p['categoria'],
                    'stock_actual': float(p['stock_actual']),
                    'stock_minimo': float(p['stock_minimo']),
                    'stock_maximo': float(p['stock_maximo']),
                    'unidad_medida': p['unidad_medida'],
                    'deficit': float(p['stock_minimo'] - p['stock_actual']),
                    'valor_stock': round(float(p['valor_stock']), 2)
                }
                for p in productos_criticos
            ],
            'productos_exceso': [
                {
                    'producto_id': p['producto_id'],
                    'codigo': p['codigo_producto'],
                    'nombre': p['nombre'],
                    'stock_actual': float(p['stock_actual']),
                    'stock_maximo': float(p['stock_maximo']),
                    'exceso': float(p['stock_actual'] - p['stock_maximo']),
                    'unidad_medida': p['unidad_medida']
                }
                for p in productos_exceso
            ],
            'productos_sin_movimiento': [
                {
                    'producto_id': p['producto_id'],
                    'codigo': p['codigo_producto'],
                    'nombre': p['nombre'],
                    'stock_actual': float(p['stock_actual']),
                    'unidad_medida': p['unidad_medida'],
                    'ultimo_movimiento': p['ultimo_movimiento'].strftime('%Y-%m-%d') if p['ultimo_movimiento'] else 'Nunca'
                }
                for p in productos_sin_movimiento
            ],
            'rotacion_promedio_dias': round(float(rotacion_promedio), 1),
            'num_criticos': len(productos_criticos),
            'num_exceso': len(productos_exceso),
            'num_sin_movimiento': len(productos_sin_movimiento)
        }
        
        # Generar según formato
        if formato == "csv":
            return generar_csv_inventario(reporte)
        elif formato == "excel":
            return generar_excel_inventario(reporte)
        elif formato == "pdf":
            pdf_buffer = generar_pdf_inventario(reporte)
            return StreamingResponse(
                pdf_buffer,
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f"attachment; filename=reporte_inventario_{datetime.now().strftime('%Y%m%d')}.pdf"
                }
            )
        else:
            return reporte
            
    finally:
        await conn.close()


def generar_csv_inventario(reporte: dict) -> StreamingResponse:
    """Genera CSV del reporte de inventario"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(['REPORTE DE ESTADO DE INVENTARIO'])
    writer.writerow([f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    writer.writerow([f"Valor Total: ${reporte['valor_total_inventario']:,.2f}"])
    writer.writerow([f"Productos: {reporte['numero_productos']}"])
    writer.writerow([f"Rotación Promedio: {reporte['rotacion_promedio_dias']} días"])
    writer.writerow([])
    
    # Productos críticos
    if reporte['productos_criticos']:
        writer.writerow(['PRODUCTOS CRÍTICOS'])
        writer.writerow(['Código', 'Nombre', 'Categoría', 'Stock', 'Mínimo', 'Déficit', 'Unidad'])
        for p in reporte['productos_criticos']:
            writer.writerow([
                p['codigo'],
                p['nombre'],
                p['categoria'],
                p['stock_actual'],
                p['stock_minimo'],
                p['deficit'],
                p['unidad_medida']
            ])
        writer.writerow([])
    
    # Productos con exceso
    if reporte['productos_exceso']:
        writer.writerow(['PRODUCTOS CON EXCESO'])
        writer.writerow(['Código', 'Nombre', 'Stock', 'Máximo', 'Exceso', 'Unidad'])
        for p in reporte['productos_exceso']:
            writer.writerow([
                p['codigo'],
                p['nombre'],
                p['stock_actual'],
                p['stock_maximo'],
                p['exceso'],
                p['unidad_medida']
            ])
    
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=reporte_inventario_{datetime.now().strftime('%Y%m%d')}.csv"
        }
    )


def generar_excel_inventario(reporte: dict) -> StreamingResponse:
    """Genera Excel del reporte de inventario"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Estado Inventario"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Título
    ws['A1'] = 'REPORTE DE ESTADO DE INVENTARIO'
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'] = f"Valor Total Inventario: ${reporte['valor_total_inventario']:,.2f}"
    ws['A4'] = f"Total Productos: {reporte['numero_productos']}"
    ws['A5'] = f"Rotación Promedio: {reporte['rotacion_promedio_dias']} días"
    
    # Productos críticos
    if reporte['productos_criticos']:
        row = 7
        ws[f'A{row}'] = f"PRODUCTOS CRÍTICOS ({reporte['num_criticos']})"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        headers = ['Código', 'Nombre', 'Categoría', 'Stock', 'Mínimo', 'Déficit', 'Unidad']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        row += 1
        for p in reporte['productos_criticos']:
            ws.cell(row=row, column=1, value=p['codigo'])
            ws.cell(row=row, column=2, value=p['nombre'])
            ws.cell(row=row, column=3, value=p['categoria'])
            ws.cell(row=row, column=4, value=p['stock_actual']).fill = warning_fill
            ws.cell(row=row, column=5, value=p['stock_minimo'])
            ws.cell(row=row, column=6, value=p['deficit'])
            ws.cell(row=row, column=7, value=p['unidad_medida'])
            row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=reporte_inventario_{datetime.now().strftime('%Y%m%d')}.xlsx"
        }
    )
